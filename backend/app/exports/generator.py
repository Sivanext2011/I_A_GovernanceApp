"""Export service: generates Excel, PDF, and PNG exports."""

import io
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from app.config import settings
from app.analytics import compute_kpis, compute_leaderboard, compute_monthly_trend
from app.utils import get_logger

logger = get_logger(__name__)

TABLE_STYLE_HEADER = TableStyle([
    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ("FONTSIZE", (0, 0), (-1, -1), 9),
    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
    ("TOPPADDING", (0, 0), (-1, -1), 4),
    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
])

TABLE_STYLE_BLUE = TableStyle([
    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E86AB")),
    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ("FONTSIZE", (0, 0), (-1, -1), 9),
    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF8FF")]),
    ("TOPPADDING", (0, 0), (-1, -1), 4),
    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
])


def _get_photo_or_none(signum: str, size=40):
    """Get photo as reportlab Image or None."""
    photo_path = settings.PHOTO_DIR / f"{signum}.jpg"
    if photo_path.exists():
        try:
            return Image(str(photo_path), width=size, height=size)
        except Exception:
            pass
    return ""


def _generate_chart_image(chart_type: str, team: str, months: list[str]) -> Path:
    """Generate a chart as PNG and return path."""
    import plotly.graph_objects as go

    trend = compute_monthly_trend(team)
    fig = go.Figure()

    if chart_type == "savings_trend":
        fig.add_trace(go.Bar(x=trend["months"], y=trend["series"]["total_savings"], name="Total", marker_color="#3b82f6"))
        fig.add_trace(go.Bar(x=trend["months"], y=trend["series"]["automation_savings"], name="Automation", marker_color="#14b8a6"))
        fig.add_trace(go.Bar(x=trend["months"], y=trend["series"]["reuse_savings"], name="Reuse", marker_color="#f59e0b"))
        fig.update_layout(title=f"Monthly Savings Trend - {team}", barmode="group")
    elif chart_type == "savings_pct":
        fig.add_trace(go.Bar(x=trend["months"], y=trend["series"]["savings_percent"], name="Savings %", marker_color="#8b5cf6"))
        fig.update_layout(title=f"Savings % Trend - {team}", yaxis_title="%")
    elif chart_type == "dept_comparison":
        from app.analytics import compute_department_comparison
        comp = compute_department_comparison(months)
        fig.add_trace(go.Bar(x=comp["teams"], y=comp["total_savings"], name="Total Savings", marker_color="#3b82f6"))
        fig.add_trace(go.Scatter(x=comp["teams"], y=comp["savings_percent"], name="Savings %", yaxis="y2", mode="lines+markers", line=dict(color="#ef4444", width=3)))
        fig.update_layout(title="Department wise Savings", yaxis2=dict(overlaying="y", side="right", title="%"), xaxis_type="category")

    fig.update_layout(template="plotly_white", font=dict(size=11), margin=dict(t=40, b=40, l=50, r=50))
    filepath = settings.EXPORT_DIR / f"_chart_{chart_type}_{team}.png"
    fig.write_image(str(filepath), width=900, height=350, scale=2)
    return filepath


def export_excel(months: list[str], team: str = "Overall") -> Path:
    from app.services.data_service import data_store
    from app.utils.departments import TEAMS

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = settings.EXPORT_DIR / f"report_{team}_{timestamp}.xlsx"

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # Overall KPI
        kpis = compute_kpis(months, team)
        pd.DataFrame([kpis]).to_excel(writer, sheet_name="KPI Summary", index=False)

        # Team-wise stats (when Overall)
        if team == "Overall":
            team_stats = []
            for t in [tt for tt in TEAMS if tt != "Overall"]:
                tk = compute_kpis(months, t)
                team_stats.append({"Team": t, **tk})
            pd.DataFrame(team_stats).to_excel(writer, sheet_name="Team Stats", index=False)

        # Leaderboard (all)
        lb = compute_leaderboard(months, team, top_n=50)
        if lb:
            pd.DataFrame(lb).to_excel(writer, sheet_name="Leaderboard", index=False)

        # Monthly Trend
        trend = compute_monthly_trend(team)
        trend_df = pd.DataFrame({"Month": trend["months"], **trend["series"]})
        trend_df.to_excel(writer, sheet_name="Monthly Trend", index=False)

        # Raw savings data
        if data_store.savings is not None:
            sav = data_store.savings
            if months:
                sav = sav[sav["CanonicalMonth"].isin(months)]
            if team != "Overall":
                sav = sav[sav["Department"] == team]
            sav.to_excel(writer, sheet_name="Savings Data", index=False)

    _style_excel(filepath)
    logger.info(f"Excel exported: {filepath}")
    return filepath


def _style_excel(filepath: Path):
    from openpyxl import load_workbook
    wb = load_workbook(filepath)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    wb.save(filepath)


def export_pdf(months: list[str], team: str = "Overall") -> Path:
    from app.services.data_service import data_store
    from app.utils.departments import TEAMS

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = settings.EXPORT_DIR / f"report_{team}_{timestamp}.pdf"

    doc = SimpleDocTemplate(str(filepath), pagesize=landscape(A4),
                            rightMargin=25, leftMargin=25, topMargin=25, bottomMargin=25)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("CustomTitle", parent=styles["Title"], fontSize=20, textColor=colors.HexColor("#1F4E79"), spaceAfter=10)
    subtitle_style = ParagraphStyle("Subtitle", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#475569"))
    heading_style = ParagraphStyle("CustomHeading", parent=styles["Heading2"], fontSize=13, textColor=colors.HexColor("#1F4E79"), spaceBefore=15, spaceAfter=8)
    elements = []

    # ===== TITLE PAGE =====
    elements.append(Spacer(1, 60))
    elements.append(Paragraph("Automation Savings Governance Report", title_style))
    elements.append(Paragraph(f"Team: {team} | Period: {months[0] if months else 'All'} to {months[-1] if months else 'All'}", subtitle_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 40))

    # ===== OVERALL KPI =====
    kpis = compute_kpis(months, team)
    elements.append(Paragraph("Overall KPI Summary", heading_style))
    kpi_data = [
        ["Total Savings", "Savings %", "Downloads", "Reused", "Pending Feedback", "Billability Hours"],
        [
            f"{kpis['total_savings']:,.2f}",
            f"{kpis['savings_percent']:.2f}%" if kpis["savings_percent"] else "N/A",
            str(kpis["total_downloads"]),
            str(kpis["total_reused_with_savings"]),
            str(kpis["pending_feedback"] or 0),
            f"{kpis['billability_hours']:,.2f}",
        ]
    ]
    t = Table(kpi_data, colWidths=[1.6 * inch] * 6)
    t.setStyle(TABLE_STYLE_HEADER)
    elements.append(t)
    elements.append(Spacer(1, 20))

    # ===== TEAM-WISE STATS (Overall only) =====
    if team == "Overall":
        elements.append(Paragraph("Team-wise Statistics", heading_style))
        team_data = [["Team", "Total Savings", "Savings %", "Downloads", "Pending", "Billability"]]
        for tt in [x for x in TEAMS if x != "Overall"]:
            tk = compute_kpis(months, tt)
            team_data.append([
                tt, f"{tk['total_savings']:,.2f}",
                f"{tk['savings_percent']:.2f}%" if tk["savings_percent"] else "N/A",
                str(tk["total_downloads"]), str(tk["pending_feedback"] or 0),
                f"{tk['billability_hours']:,.2f}",
            ])
        t = Table(team_data, colWidths=[1.8 * inch, 1.5 * inch, 1.2 * inch, 1.2 * inch, 1.0 * inch, 1.5 * inch])
        t.setStyle(TABLE_STYLE_HEADER)
        elements.append(t)
        elements.append(Spacer(1, 20))

    # ===== CURRENT MONTH =====
    latest = data_store.get_latest_month()
    if latest:
        elements.append(Paragraph(f"Current Month: {latest}", heading_style))
        curr = compute_kpis([latest], team)
        curr_data = [
            ["Monthly Savings", "Monthly Savings %", "Monthly Downloads", "Monthly Billability"],
            [
                f"{curr['total_savings']:,.2f}",
                f"{curr['savings_percent']:.2f}%" if curr["savings_percent"] else "N/A",
                str(curr["total_downloads"]),
                f"{curr['billability_hours']:,.2f}",
            ]
        ]
        t = Table(curr_data, colWidths=[2.2 * inch] * 4)
        t.setStyle(TABLE_STYLE_BLUE)
        elements.append(t)

    elements.append(PageBreak())

    # ===== CHARTS =====
    elements.append(Paragraph("Monthly Savings Trend", heading_style))
    try:
        chart_path = _generate_chart_image("savings_trend", team, months)
        elements.append(Image(str(chart_path), width=9 * inch, height=3.2 * inch))
    except Exception as e:
        elements.append(Paragraph(f"Chart generation failed: {e}", styles["Normal"]))
    elements.append(Spacer(1, 15))

    elements.append(Paragraph("Savings % Trend", heading_style))
    try:
        chart_path = _generate_chart_image("savings_pct", team, months)
        elements.append(Image(str(chart_path), width=9 * inch, height=3.2 * inch))
    except Exception as e:
        elements.append(Paragraph(f"Chart generation failed: {e}", styles["Normal"]))

    if team == "Overall":
        elements.append(PageBreak())
        elements.append(Paragraph("Department wise Savings", heading_style))
        try:
            chart_path = _generate_chart_image("dept_comparison", team, months)
            elements.append(Image(str(chart_path), width=9 * inch, height=3.2 * inch))
        except Exception as e:
            elements.append(Paragraph(f"Chart generation failed: {e}", styles["Normal"]))

    elements.append(PageBreak())

    # ===== MONTHLY TREND TABLE =====
    trend = compute_monthly_trend(team)
    if trend["months"]:
        elements.append(Paragraph("Monthly Trend Data", heading_style))
        trend_data = [["Month", "Total Savings", "Automation", "Reuse", "Savings %"]]
        for i, m in enumerate(trend["months"]):
            trend_data.append([
                m, f"{trend['series']['total_savings'][i]:,.2f}",
                f"{trend['series']['automation_savings'][i]:,.2f}",
                f"{trend['series']['reuse_savings'][i]:,.2f}",
                f"{trend['series']['savings_percent'][i]:.2f}%",
            ])
        t = Table(trend_data, colWidths=[1.5 * inch, 1.8 * inch, 1.8 * inch, 1.8 * inch, 1.5 * inch])
        t.setStyle(TABLE_STYLE_HEADER)
        elements.append(t)

    elements.append(PageBreak())

    # ===== LEADERBOARD WITH PHOTOS =====
    lb = compute_leaderboard(months, team, top_n=30)
    if lb:
        elements.append(Paragraph("Leaderboard - Top Practitioners", heading_style))
        lb_data = [["#", "Photo", "Name", "Department", "Reuse", "Automation", "Total Savings"]]
        for i, entry in enumerate(lb, 1):
            photo = _get_photo_or_none(entry["signum"], size=30)
            lb_data.append([
                str(i), photo, entry["name"], entry["department"],
                f"{entry['reuse_saving']:,.2f}", f"{entry['automation_saving']:,.2f}",
                f"{entry['total_savings']:,.2f}"
            ])

        col_widths = [0.4 * inch, 0.6 * inch, 2.0 * inch, 1.6 * inch, 1.3 * inch, 1.3 * inch, 1.3 * inch]
        t = Table(lb_data, colWidths=col_widths, rowHeights=[20] + [35] * len(lb))
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E86AB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF8FF")]),
        ]))
        elements.append(t)

    doc.build(elements)

    # Cleanup temp chart images
    for f in settings.EXPORT_DIR.glob("_chart_*.png"):
        f.unlink(missing_ok=True)

    logger.info(f"PDF exported: {filepath}")
    return filepath


def export_png_chart(months: list[str], team: str = "Overall", chart_type: str = "trend") -> Path:
    import plotly.graph_objects as go

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = settings.EXPORT_DIR / f"chart_{chart_type}_{team}_{timestamp}.png"

    trend = compute_monthly_trend(team)
    fig = go.Figure()

    if chart_type == "trend":
        fig.add_trace(go.Bar(x=trend["months"], y=trend["series"]["total_savings"], name="Total Savings", marker_color="#1F4E79"))
        fig.add_trace(go.Bar(x=trend["months"], y=trend["series"]["automation_savings"], name="Automation", marker_color="#14b8a6"))
        fig.add_trace(go.Bar(x=trend["months"], y=trend["series"]["reuse_savings"], name="Reuse", marker_color="#f59e0b"))
        fig.update_layout(title=f"Monthly Savings Trend - {team}", barmode="group")
    elif chart_type == "savings_pct":
        fig.add_trace(go.Bar(x=trend["months"], y=trend["series"]["savings_percent"], name="Savings %", marker_color="#8b5cf6"))
        fig.update_layout(title=f"Savings % Trend - {team}", yaxis_title="%")

    fig.update_layout(template="plotly_white", font=dict(size=12))
    fig.write_image(str(filepath), width=1200, height=600, scale=2)
    logger.info(f"PNG exported: {filepath}")
    return filepath
